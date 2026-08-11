import argparse
import datetime
import os
from collections.abc import Callable
from logging import INFO

import pandas as pd
from azure.data.tables import TableClient
from azure.identity import DefaultAzureCredential
from dotenv import load_dotenv
from nhpy.az import connect_to_container, load_parquet_file
from nhpy.utils import configure_logging, get_logger
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from nhp.capacity_conversion.config import ASSUMPTIONS_URL

logger = get_logger()

# Suppression methodology follows NHS England HES standard:
# https://digital.nhs.uk/data-and-information/publications/statistical/
# hospital-admitted-patient-care-activity/supporting-information#suppression-methodology
SUPPRESSION_THRESHOLD = 7  # suppress counts 1–7; round all others to nearest 5


def get_baseline_activity(aggregations: pd.DataFrame) -> pd.DataFrame:
    """Extract baseline (model run 0) total activity per functional area.

    Applies NHS England HES suppression rules:
    - Values 1–7 are replaced with None (displayed as blank in Excel)
    - All other non-zero values are rounded to the nearest 5

    Args:
        aggregations (pd.DataFrame): Raw aggregations with model_run index,
            grouping and total columns

    Returns:
        pd.DataFrame: Baseline activity per grouping, suppressed and rounded
    """
    value_columns = aggregations.select_dtypes("number").columns.tolist()
    baseline = (
        aggregations.loc[aggregations.index == 0]
        .groupby("grouping")[value_columns]
        .sum()
    )

    def _suppress_and_round(x: float) -> float | None:
        if 1 <= x <= SUPPRESSION_THRESHOLD:
            return None
        return round(x / 5) * 5

    return baseline.map(_suppress_and_round)


def calculate_prediction_intervals_and_mean(
    activity_column: pd.Series,
) -> dict[str, float]:
    """Calculate p10, p90 and mean for activity in each functional area

    Args:
        activity_column (pd.Series): Column with activity counts for each functional area

    Returns:
        dict[str, float]: Dictionary with p10, p90 and mean as keys
    """
    results_dict = {"mean": float(activity_column.mean())}
    results_dict["p10"] = float(activity_column.quantile(0.1))
    results_dict["p90"] = float(activity_column.quantile(0.9))
    return results_dict


def summarise_model_runs(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate p10, p90 and mean across all model runs

    Args:
        df (pd.DataFrame): MultiIndex DataFrame with one index called "model_run"

    Raises:
        ValueError: If more than one value column in DataFrame

    Returns:
        pd.DataFrame: Summarised DataFrame
    """
    group_col_names = [name for name in df.index.names if name != "model_run"]
    if len(group_col_names) != 1:
        raise ValueError("Expected exactly one index column.")
    value_cols = [c for c in df.columns if c != "model_run"]
    if len(value_cols) > 1:
        df_list = []
        for col in value_cols:
            summary_df = pd.DataFrame(
                df.groupby(level=group_col_names)[col].agg(
                    p10=lambda s: s.quantile(0.10),
                    mean="mean",
                    p90=lambda s: s.quantile(0.90),
                )
            )
            summary_df["measure"] = col
            df_list.append(summary_df.reset_index())
        return pd.concat(df_list).set_index(["grouping", "measure"]).sort_index()
    return pd.DataFrame(
        df.groupby(level=group_col_names)[value_cols[0]].agg(
            p10=lambda s: s.quantile(0.10),
            mean="mean",
            p90=lambda s: s.quantile(0.90),
        )
    )


def load_assumptions(path_to_csv: str) -> pd.DataFrame:
    """Loads assumptions for use in model. Defaults to assumptions published online at
    https://the-strategy-unit.github.io/open-plan-docs/reference/functional-area-catalogue/

    Args:
        path_to_csv (str): Path to assumptions csv file

    Returns:
        pd.DataFrame: Dataframe with assumption values and variable names
    """
    logger.info(f"Loading assumptions from {path_to_csv}...")
    return pd.read_csv(path_to_csv).set_index("Assumption ID")[["Value"]].sort_index()


def process_and_save_results_to_excel(
    data_to_save: dict[str, pd.DataFrame | pd.Series],
) -> None:
    """Saves results of capacity conversion pipeline to Excel

    Args:
        data_to_save (dict[str, pd.DataFrame  |  pd.Series]): Dictionary of data to save, where the keys are the titles of the
        worksheets and the values are the dataframes to be included. At minimum should include "metadata" key and dataframe.
    """
    directory = os.path.join(
        "results",
        str(data_to_save["metadata"].loc["guid"]),
        str(data_to_save["metadata"].loc["capacity_conversion_runtime"]),
    )
    os.makedirs(directory, exist_ok=True)
    filepath = os.path.join(directory, "capacity_conversion_results.xlsx")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for sheet_name, df in data_to_save.items():
        if isinstance(df, pd.DataFrame) and "model_run" in df.index.names:
            df = summarise_model_runs(df)
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(
            dataframe_to_rows(pd.DataFrame(df).reset_index(), index=False, header=True),
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
    wb.save(filepath)
    logger.info(f"💾 Results saved to {filepath}")


def load_metadata_from_ats(
    guid: str,
    storage_endpoint: str,
    table_name: str,
    capacity_model_version: str,
) -> dict:
    """Loads metadata for scenario converted to functional area aggregations
    from Azure Table Storage

    Args:
        guid (str): GUID for functional area aggregation
        storage_endpoint (str): Azure Table Storage endpoint, in format "https://{storage_account_name}.table.core.windows.net"
        table_name (str): Table name containing metadata for Functional Area Aggregations
        capacity_model_version (str): Version of capacity model.

    Returns:
        dict: Dictionary with metadata for given Functional Area aggregation
    """
    credential = DefaultAzureCredential()
    table_client = TableClient(
        endpoint=storage_endpoint, table_name=table_name, credential=credential
    )
    entity = table_client.get_entity(partition_key=capacity_model_version, row_key=guid)
    metadata = dict(entity)
    metadata["guid"] = guid
    metadata["capacity_model_version"] = capacity_model_version
    return metadata


def create_aggregations_path(metadata: dict) -> str:
    """Create path to aggregations parquet files on Azure Storage

    Args:
        metadata (dict): Dictionary of metadata for capacity conversion

    Returns:
        str: Full path to the specific functional area aggregations to be converted to capacity
    """
    return f"functional-aggregations/{metadata['capacity_model_version']}/{metadata['guid']}/"


def validate_required_env_vars() -> dict:
    """
    Loads environment variables and ensures required variables are present.
    Raises EnvironmentError if any are missing or empty.
    Returns a dictionary of the validated variables.
    """

    load_dotenv()

    required_vars = [
        "AZ_STORAGE_EP",
        "AZ_STORAGE_RESULTS",
        "TABLE_NAME",
        "AZ_TABLE_ENDPOINT",
    ]

    values = {}
    missing = []

    for var in required_vars:
        value = os.getenv(var)
        if not value:
            missing.append(var)
        else:
            values[var] = value

    if missing:
        raise OSError(
            f"Missing required environment variables in .env: {', '.join(missing)}"
        )

    return values


def load_aggregations(
    account_url: str,
    results_container: str,
    aggregations_path: str,
    aggregation_type: str,
) -> pd.DataFrame:
    """Loads aggregated data from Azure

    Args:
        account_url (str): Azure Storage account URL
        results_container (str): Azure Storage container name with results
        aggregations_path (str): Path to "folder" with data to load
        aggregation_type (str): Path to

    Returns:
        pd.DataFrame: Loads aggregated data
    """
    logger.info(f"Loading {aggregation_type} data from {aggregations_path}...")
    results_connection = connect_to_container(account_url, results_container)
    aggregations = load_parquet_file(
        results_connection,
        f"{aggregations_path.rstrip('/')}/{aggregation_type}.parquet",
    )
    return aggregations


def process_activity_type(
    name: str,
    aggregations: pd.DataFrame,
    calculate_fn: Callable,
    assumptions: pd.DataFrame,
    data_to_save: dict[str, pd.DataFrame | pd.Series],
    preprocess: Callable[[pd.DataFrame], pd.DataFrame] | None = None,
    include_baseline: bool = True,
) -> None:
    """Summarise functional areas, optionally extract baseline, and calculate capacity."""
    if preprocess is not None:
        aggregations = preprocess(aggregations)
    # We exclude baseline (model run 0) from conversion to capacity
    functional_areas = (
        aggregations[aggregations.index != 0]
        .reset_index()
        .set_index(["grouping", "model_run"])
    )
    data_to_save[f"{name}_fun_area_groupings"] = functional_areas
    if include_baseline:
        data_to_save[f"{name}_baseline"] = get_baseline_activity(aggregations)
    capacity_df = calculate_fn(functional_areas, assumptions)
    data_to_save[f"{name}_capacity"] = capacity_df


def run_single_activity_type(
    activity_type: str,
    calculate_fn: Callable,
    preprocess: Callable | None = None,
    include_baseline: bool = False,
) -> int:
    """CLI entry point for a single activity type.

    Handles argument parsing, metadata/assumptions loading, aggregation loading,
    optional preprocessing, capacity calculation, and Excel saving.
    """
    configure_logging(INFO)
    capacity_conversion_runtime = datetime.datetime.now(tz=datetime.UTC).strftime(
        "%Y%m%d_%H%M%S"
    )

    parser = argparse.ArgumentParser(
        description=f"Generate {activity_type.upper()} capacity outputs given functional area aggregations of {activity_type.upper()} activity"
    )
    parser.add_argument(
        "guid",
        help="GUID of functional area aggregation to convert into capacity",
    )
    parser.add_argument(
        "--capacity_model_version",
        help="Capacity model version",
        default="dev",
    )
    parser.add_argument(
        "--path_to_assumptions_file",
        help=f"Path to assumptions file (default: '{ASSUMPTIONS_URL}')",
        default=ASSUMPTIONS_URL,
    )
    args = parser.parse_args()

    config = validate_required_env_vars()
    data_to_save = {}

    metadata = load_metadata_from_ats(
        args.guid,
        config["AZ_TABLE_ENDPOINT"],
        config["TABLE_NAME"],
        args.capacity_model_version,
    )
    metadata["capacity_conversion_runtime"] = capacity_conversion_runtime
    data_to_save["metadata"] = pd.Series(metadata).drop(["PartitionKey", "RowKey"])

    assumptions = load_assumptions(args.path_to_assumptions_file)
    data_to_save["assumptions"] = assumptions

    aggregations_path = create_aggregations_path(metadata)
    aggregations = load_aggregations(
        config["AZ_STORAGE_EP"],
        config["AZ_STORAGE_RESULTS"],
        aggregations_path,
        activity_type,
    )

    process_activity_type(
        name=activity_type,
        aggregations=aggregations,
        calculate_fn=calculate_fn,
        assumptions=assumptions,
        data_to_save=data_to_save,
        preprocess=preprocess,
        include_baseline=include_baseline,
    )

    process_and_save_results_to_excel(data_to_save)
    return 0
