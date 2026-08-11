import importlib.util
import os
from io import BytesIO
from pathlib import Path
from types import ModuleType
from unittest.mock import call

import pandas as pd
import pytest
from openpyxl import load_workbook


def _load_app_module() -> ModuleType:
    app_path = Path(__file__).parents[2] / "app.py"
    spec = importlib.util.spec_from_file_location("capacity_conversion_app", app_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to load {app_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


app = _load_app_module()

FUNCTIONAL_AGGREGATION_ENVIRONMENT = {
    "AZ_FUNC_AGG_GUID": "guid-123",
    "AZ_STORAGE_EP": "https://storage.example.com",
    "AZ_STORAGE_RESULTS": "results",
    "AZ_TABLE_ENDPOINT": "https://table.example.com",
    "TABLE_NAME": "metadata",
}


def test_load_capacity_results(mocker):
    mocker.patch.dict(
        os.environ,
        FUNCTIONAL_AGGREGATION_ENVIRONMENT,
        clear=True,
    )
    mock_datetime = mocker.patch.object(app, "datetime")
    mock_datetime.now.return_value.strftime.return_value = "20260101_120000"

    metadata = {
        "PartitionKey": "dev",
        "RowKey": "guid-123",
        "guid": "guid-123",
        "capacity_model_version": "dev",
    }
    load_metadata = mocker.patch.object(
        app,
        "load_metadata_from_ats",
        return_value=metadata,
    )
    create_path = mocker.patch.object(
        app,
        "create_aggregations_path",
        return_value="functional-aggregations/dev/guid-123/",
    )
    aggregations = pd.DataFrame({"total": [1]})
    load_aggregation = mocker.patch.object(
        app,
        "load_aggregations",
        return_value=aggregations,
    )
    assumptions = pd.DataFrame({"Value": [1]})
    mocker.patch.object(app, "load_assumptions", return_value=assumptions)
    process = mocker.patch.object(app, "process_activity_type")

    data_to_save = app._load_capacity_results()

    load_metadata.assert_called_once_with(
        "guid-123",
        "https://table.example.com",
        "metadata",
        "dev",
    )
    create_path.assert_called_once_with(metadata)
    load_aggregation.assert_has_calls(
        [
            call(
                "https://storage.example.com",
                "results",
                "functional-aggregations/dev/guid-123/",
                activity_type,
            )
            for activity_type in app.ACTIVITY_TYPES
        ]
    )
    assert load_aggregation.call_count == 4
    process.assert_has_calls(
        [
            call(
                "op",
                aggregations,
                app.calculate_op_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "aae",
                aggregations,
                app.calculate_aae_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "ip_daycase",
                aggregations,
                app.calculate_daycase_capacity,
                assumptions,
                data_to_save,
                preprocess=None,
            ),
            call(
                "ip_maternity",
                aggregations,
                app.calculate_maternity_capacity,
                assumptions,
                data_to_save,
                preprocess=app.preprocess_ip_maternity_data,
            ),
        ]
    )
    assert process.call_count == 4
    mock_datetime.now.assert_called_once_with(tz=app.UTC)
    assert data_to_save["metadata"].to_dict() == {
        "guid": "guid-123",
        "capacity_model_version": "dev",
        "capacity_conversion_runtime": "20260101_120000",
    }


def test_load_capacity_results_requires_guid(mocker):
    mocker.patch.dict(
        os.environ,
        {
            key: value
            for key, value in FUNCTIONAL_AGGREGATION_ENVIRONMENT.items()
            if key != "AZ_FUNC_AGG_GUID"
        },
        clear=True,
    )

    with pytest.raises(
        RuntimeError,
        match="Missing required environment variable: AZ_FUNC_AGG_GUID",
    ):
        app._load_capacity_results()


def test_create_workbook():
    capacity = pd.DataFrame(
        {
            "output": ["room", "room"],
            "model_run": [1, 2],
            "value": [1.0, 3.0],
        }
    ).set_index(["output", "model_run"])
    data_to_save = {
        "metadata": pd.Series({"guid": "guid-123"}),
        "op_capacity": capacity,
    }

    workbook = load_workbook(BytesIO(app._create_workbook(data_to_save)))

    assert workbook.sheetnames == ["metadata", "op_capacity"]
    rows = list(workbook["op_capacity"].values)
    assert rows[0] == ("output", "p10", "mean", "p90")
    assert rows[1] == ("room", 1.2, 2, 2.8)
